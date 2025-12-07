import openpyxl
from pathlib import Path


class SealCheckConverterCore:
    """Core conversion functionality"""
    
    def __init__(self, template_path):
        self.template_path = Path(template_path)
        
    def find_column_index(self, headers, column_name):
        """Find the index of a column by name (case-insensitive)"""
        for i, header in enumerate(headers, 1):
            if header and str(header).strip().upper() == column_name.upper():
                return i
        return None
    
    def get_non_empty_value(self, cell_value):
        """Return cell value only if it's not empty"""
        if cell_value is None:
            return None
        
        # Handle NaN from pandas/Excel
        if isinstance(cell_value, float):
            import math
            if math.isnan(cell_value):
                return None
        
        # Handle string "NaN" or "nan"
        if isinstance(cell_value, str):
            cell_value = cell_value.strip()
            if cell_value.upper() in ['', 'NAN', ',----------------------', '0']:
                return None
        
        # Handle numeric 0
        if cell_value == 0:
            return None
            
        return cell_value
    
    def get_pol_value(self, ws):
        """Get POL value from first data row for naming purposes"""
        try:
            pol = ws.cell(row=2, column=2).value
            if pol:
                return str(pol).strip().upper()
        except:
            pass
        return None
    
    def detect_file_format(self, ws_source):
        """Detect file format: GATE_IN, LISTADO, RODMAN, UNITLIST, COLON YARD, or PISCO"""
        headers_row1 = [cell.value for cell in ws_source[1]]
        headers_row3 = [cell.value for cell in ws_source[3]]
        
        # Check for PISCO format (has "CtrNbr" and "POR" in row 1)
        pisco_indicators = ['CtrNbr', 'POR', 'SzTp']
        if all(h in headers_row1 for h in pisco_indicators):
            return 'PISCO', headers_row1, 2
        
        # Check for LIST_OF_UNIT format (has "UNIT", "SIZE", "POL", "POD", "SEAL")
        listofunit_indicators = ['UNIT', 'SIZE', 'POL', 'POD', 'SEAL']
        if all(h in headers_row1 for h in listofunit_indicators):
            return 'LIST_OF_UNIT', headers_row1, 2
        
        # Check for UNITLIST format (has "Voyage In" column)  
        unitlist_indicators = ['Voyage In', 'Voyage Out']
        if any(h and h in unitlist_indicators for h in headers_row1):
            return 'UNITLIST', headers_row1, 2
        
        # Check for Colon Yard format (has "Current LOC Block" column)  
        colonlist_indicators = ['Current LOC Block', 'Current LOC Bay']
        if any(h and h in colonlist_indicators for h in headers_row1):
            return 'COLON YARD', headers_row1, 9
        
        # Check for RODMAN format (has "Slot (Yard)" in row 1)
        if any(h and 'Slot (Yard)' in str(h) for h in headers_row1):
            # Check if already converted
            if 'Container' == headers_row1[0]:
                return 'RODMAN_CONVERTED', headers_row1, 2
            return 'RODMAN', headers_row1, 2
        
        # Check for LISTADO format (row 3 has CONTENEDOR)
        if any(h and 'CONTENEDOR' in str(h).upper() for h in headers_row3):
            return 'LISTADO', headers_row3, 4
        
        # Default to GATE_IN format
        return 'GATE_IN', headers_row1, 2
    
    def convert(self, source_path, output_path, progress_callback=None, voyage_filters=None, carrier_filters=None):
        """
        Convert source file to seal check template format
        
        Args:
            source_path: Path to source Excel file
            output_path: Path for output file
            progress_callback: Optional callback function(message) for progress updates
            voyage_filters: Optional list of voyage numbers to filter (for UnitList format)
            
        Returns:
            dict with conversion results or None on error
        """
        def log(message):
            if progress_callback:
                progress_callback(message)
        
        try:
            log(f"\n🚀 STARTING CONVERSION")
            log(f"📂 Source: {Path(source_path).name}")
            log(f"📋 Template: {self.template_path.name}")
            
            # Load template
            log("⏳ Loading template...")
            wb_template = openpyxl.load_workbook(self.template_path)
            ws_template = wb_template.active
            
            # Load source
            log("⏳ Loading source...")
            wb_source = openpyxl.load_workbook(source_path, data_only=True)
            ws_source = wb_source.active
            
            # Detect format
            file_format, headers, data_start_row = self.detect_file_format(ws_source)
            log(f"✓ Detected {file_format} format (headers in row {data_start_row-1})")
            
            # Column mappings - handle all formats
            if file_format in ['RODMAN', 'RODMAN_CONVERTED']:
                # Rodman format mappings
                direct_mappings = {
                    'Unit': 'Container',
                    'Container': 'Container',
                    'POL': 'POL',
                    'POD': 'POD',
                    'ISO': 'Type',
                    'Type': 'Type',
                    'Slot (Yard)': 'Slot',
                    'Slot': 'Slot'
                }
                seal_columns = [
                    'Seal 1 / type / origin',
                    'Seal 2 / type / origin',
                    'Seal 3 / type / origin',
                    'Seal 1', 'Seal 2', 'Seal 3'
                ]
            elif file_format == 'PISCO':
                # Pisco Load List format mappings
                direct_mappings = {
                    'CtrNbr': 'Container',
                    'POR': 'POL',
                    'POD': 'POD',
                    'SzTp': 'Type'
                }
                seal_columns = [
                    'Carrier Seal'
                ]
            elif file_format == 'UNITLIST':
                # UnitList format mappings
                direct_mappings = {
                    'Unit': 'Container',
                    'POL': 'POL',
                    'SPOD': 'POD',
                    'ISO': 'Type',
                    'Slot (EXE)': 'Slot',
                    'Voyage In': None,  # Not mapped to template, but needed for filtering
                    'Voyage Out': None  # Not mapped to template, but needed for filtering
                }
                # UnitList seal columns - extract seal before " / "
                seal_columns = [
                    'Seal 1 / type / origin',
                    'Seal 2 / type / origin',
                    'Seal 3 / type / origin'
                ]
            elif file_format == 'COLON YARD':
                # COLON YARD format mappings
                direct_mappings = {
                    'Container No': 'Container',
                    'POL': 'POL',
                    'POD': 'POD',
                    'ISO Type': 'Type',
                    # Slot needs special handling - will merge Block, Bay, Row, Tier
                    'Dept Carrier': None,  # Not mapped to template, but needed for filtering
                    'Carrier': None,  # Not mapped to template, but needed for MSC filtering
                }
                # COLON YARD seal columns
                seal_columns = [
                    'Seal No. 1'
                ]
            elif file_format == 'LIST_OF_UNIT':
                # LIST_OF_UNIT format mappings
                direct_mappings = {
                    'UNIT': 'Container',
                    'SIZE': 'Type',
                    'POL': 'POL',
                    'POD': 'POD'
                }
                # Handle multiple SEAL columns with same header name
                seal_columns = []
                # Find all columns named "SEAL" by checking each column index
                for col_idx, header in enumerate(headers, 1):
                    if header and str(header).upper() == 'SEAL':
                        seal_columns.append(f'SEAL_COL_{col_idx}')  # Create unique identifiers
            elif file_format == 'LISTADO':
                # LISTADO format mappings
                direct_mappings = {
                    'CONTENEDOR': 'Container',
                    'POT (PUERTO DE DESCARGA)': 'POL',  # Will be overridden to ECGYE
                    'POD (PUERTO DE DESTINO FINAL)': 'POD',
                    'POD (PUERTO FINAL)': 'POD',
                    'SIZE': 'Type'
                }
                seal_columns = [
                    'NAVIERO', 
                    'Seal 2(shipping)', 'SEAL 2',
                    'Seal EL (shipping)', 'SELLO CABLE (EL)',
                    'OTHER SEAL (shipping)', 'SEAL4 (XZ)', 'SEAL 5 (DL)',
                    'SEAL 6', 'SEAL 7', 'SEAL 8', 'SEAL 9', 'SEAL 10',
                    'Seal (PAN)', 'SelloPAN',
                    'Seal (AFORO)', 'Sello AFORO', 'AFORO',
                    'Seal (2)', 'Sello2',
                    'Other Seal New',
                    'Naviero', 'Sello3', 'Sello4', 'Sello5', 'Sello'
                ]
            else:
                # GATE_IN format (default)
                direct_mappings = {
                    'CONTENEDOR': 'Container',
                    'PortLoad': 'POL',  # Use PortLoad for GATE_IN
                    'POD (PUERTO DE DESTINO FINAL)': 'POD',
                    'POD (PUERTO FINAL)': 'POD',
                    'SIZE': 'Type'
                }
                seal_columns = [
                    'NAVIERO', 
                    'Seal 2(shipping)', 'SEAL 2',
                    'Seal EL (shipping)', 'SELLO CABLE (EL)',
                    'OTHER SEAL (shipping)', 'SEAL4 (XZ)', 'SEAL 5 (DL)',
                    'SEAL 6', 'SEAL 7', 'SEAL 8', 'SEAL 9', 'SEAL 10',
                    'Seal (PAN)', 'SelloPAN',
                    'Seal (AFORO)', 'Sello AFORO', 'AFORO',
                    'Seal (2)', 'Sello2',
                    'Other Seal New',
                    'Naviero', 'Sello3', 'Sello4', 'Sello5', 'Sello'
                ]
            
            # Find column indices
            source_col_indices = {}
            columns_to_find = list(direct_mappings.keys()) + seal_columns
            
            # Add slot component columns for COLON YARD
            if file_format == 'COLON YARD':
                columns_to_find.extend(['Current LOC Block', 'Current LOC Bay', 'Current LOC Row', 'Current LOC Tier'])
            # Special handling for LIST_OF_UNIT: Find all SEAL columns by index
            if file_format == 'LIST_OF_UNIT':
                seal_col_indices = []
                for col_idx, header in enumerate(headers, 1):
                    if header and str(header).upper() == 'SEAL':
                        seal_col_indices.append(col_idx)
                        source_col_indices[f'SEAL_COL_{col_idx}'] = col_idx
                columns_to_find = [c for c in columns_to_find if not c.startswith('SEAL_COL_')]  # Remove seal placeholders
            for col_name in columns_to_find:
                idx = self.find_column_index(headers, col_name)
                if idx:
                    source_col_indices[col_name] = idx
            
            log(f"✓ Found {len(source_col_indices)} columns")
            
            # Determine container column name based on format
            if file_format in ['RODMAN', 'RODMAN_CONVERTED']:
                container_col_name = 'Unit' if 'Unit' in source_col_indices else 'Container'
            elif file_format == 'UNITLIST':
                container_col_name = 'Unit'
            elif file_format == 'PISCO':
                container_col_name = 'CtrNbr'
            elif file_format == 'LIST_OF_UNIT':
                container_col_name = 'UNIT'
            elif file_format == 'COLON YARD':
                container_col_name = 'Container No'
            else:
                container_col_name = 'CONTENEDOR'
            
            # Check for container column
            if container_col_name not in source_col_indices:
                log(f"❌ ERROR: {container_col_name} column not found!")
                return None
            
            container_col_idx = source_col_indices[container_col_name]
            
            # Get voyage column index for filtering (UnitList only)
            voyage_col_idx = None
            if file_format == 'UNITLIST' and voyage_filters:
                if 'Voyage Out' in source_col_indices:
                    voyage_col_idx = source_col_indices['Voyage Out']
                    log(f"✓ Filtering by Voyage Out: {', '.join(voyage_filters)}")
                elif 'Voyage In' in source_col_indices: # might need delete
                    voyage_col_idx = source_col_indices['Voyage In']
                    log(f"✓ Filtering by Voyage In: {', '.join(voyage_filters)}")

            # Get carrier column index for filtering (COLON YARD only)
            carrier_col_idx = None
            if file_format == 'COLON YARD' and carrier_filters:
                if 'Dept Carrier' in source_col_indices:
                    carrier_col_idx = source_col_indices['Dept Carrier']
                    log(f"✓ Filtering by Dept Carrier: {', '.join(carrier_filters)}")

            if file_format not in ['UNITLIST', 'COLON YARD'] and (voyage_filters or carrier_filters):
                # Ignore voyage filters for non-UnitList formats
                log(f"ℹ️  Voyage filters ignored for {file_format} format")
            
            # Template columns
            template_cols = {
                'Container': 1, 'POL': 2, 'POD': 3, 'Type': 4, 'Slot': 5,
                'Seal 1': 6, 'Seal 2': 7, 'Seal 3': 8, 'Seal 4': 9, 'Seal 5': 10
            }
            
            # Statistics
            containers_found = 0
            containers_skipped = 0
            total_seals = 0
            seal_distribution = {}
            
            output_row = 2
            
            log("⚙️  Processing rows...")
            if voyage_col_idx:
                log(f"   (Filtering by Voyage column)")
            elif carrier_col_idx:
                log(f"   (Filtering by Dept Carrier column)")
            else:
                log(f"   (Stops at first empty {container_col_name})")
            
            # Process each row
            for source_row in range(data_start_row, ws_source.max_row + 1):
                # Check if container has a value
                container_value = ws_source.cell(row=source_row, column=container_col_idx).value
                container_value = self.get_non_empty_value(container_value)
                
                # STOP if empty container found (except when filtering)
                if container_value is None:
                    if not voyage_col_idx and not carrier_col_idx:
                        log(f"⏹️  Stopped at row {source_row} (empty {container_col_name})")
                        break
                    else:
                        continue  # Skip empty rows when filtering

                # Filter by Voyage In if specified (UnitList only)
                if voyage_col_idx and voyage_filters:
                    voyage_value = ws_source.cell(row=source_row, column=voyage_col_idx).value
                    voyage_value = self.get_non_empty_value(voyage_value)
                    


                    # Skip if voyage doesn't match any filter
                    if voyage_value not in voyage_filters:
                        containers_skipped += 1
                        continue

                # Filter by Dept Carrier if specified (COLON YARD only)
                if carrier_col_idx and carrier_filters:
                    carrier_value = ws_source.cell(row=source_row, column=carrier_col_idx).value
                    carrier_value = self.get_non_empty_value(carrier_value)
                    
                    # Skip rows with empty/None carrier when filtering
                    if not carrier_value:
                        containers_skipped += 1
                        continue
                    
                    # Check if carrier matches any filter (supports partial matching)
                    carrier_value_upper = str(carrier_value).upper()
                    matches_filter = False
                    for filter_carrier in carrier_filters:
                        filter_carrier_upper = str(filter_carrier).upper()
                        if filter_carrier_upper in carrier_value_upper:
                            matches_filter = True
                            break
                    
                    if not matches_filter:
                        containers_skipped += 1
                        continue

                # Additional MSC carrier check for COLON YARD
                if file_format == 'COLON YARD' and 'Carrier' in source_col_indices:
                    msc_carrier_col_idx = source_col_indices['Carrier']
                    msc_carrier_value = ws_source.cell(row=source_row, column=msc_carrier_col_idx).value
                    msc_carrier_value = self.get_non_empty_value(msc_carrier_value)
                    
                    # Only process containers where Carrier = 'MSC'
                    if not msc_carrier_value or str(msc_carrier_value).upper() != 'MSC':
                        containers_skipped += 1
                        continue

                containers_found += 1
                
                # Copy direct mappings
                for source_col, template_col_name in direct_mappings.items():
                    # Skip if no template mapping (e.g., Voyage columns for filtering only)
                    if template_col_name is None:
                        continue
                        
                    if source_col in source_col_indices:
                        source_col_idx = source_col_indices[source_col]
                        template_col_idx = template_cols[template_col_name]
                        
                        value = ws_source.cell(row=source_row, column=source_col_idx).value
                        value = self.get_non_empty_value(value)
                        
                        # Special case: POL defaults based on format
                        if template_col_name == 'POL':
                            if file_format == 'UNITLIST':
                                # UnitList always uses PSA-RODMAN
                                value = 'PSA-RODMAN'
                            elif file_format == 'LISTADO':
                                # LISTADO always uses ECGYE (override source value)
                                value = 'ECGYE'
                            elif file_format in ['RODMAN', 'RODMAN_CONVERTED'] and value is None:
                                # RODMAN uses RODMAN only if empty
                                value = 'RODMAN'
                        
                        # Write the value (None for empty cells)
                        if value is not None:
                            ws_template.cell(row=output_row, column=template_col_idx).value = value
                
                # Special case: Merge Slot columns for COLON YARD
                if file_format == 'COLON YARD':
                    slot_parts = []
                    # Get the slot component columns
                    slot_columns = ['Current LOC Block', 'Current LOC Bay', 'Current LOC Row', 'Current LOC Tier']
                    for slot_col in slot_columns:
                        if slot_col in source_col_indices:
                            slot_col_idx = source_col_indices[slot_col]
                            slot_value = ws_source.cell(row=source_row, column=slot_col_idx).value
                            slot_value = self.get_non_empty_value(slot_value)
                            if slot_value:
                                # Convert to string and pad bay/row/tier with zeros if needed
                                slot_value_str = str(slot_value)
                                if slot_col in ['Current LOC Bay', 'Current LOC Row']:
                                    slot_value_str = slot_value_str.zfill(2)  # Pad to 2 digits
                                slot_parts.append(slot_value_str)
                    
                    # Merge the parts and write to Slot column
                    if slot_parts:
                        merged_slot = ''.join(slot_parts)
                        template_col_idx = template_cols['Slot']
                        ws_template.cell(row=output_row, column=template_col_idx).value = merged_slot
                
                # Process seals - avoid duplicates, combine extras in Seal 5
                seals_added = set()
                seals_list = []
                
                # Collect all unique seals
                for seal_col in seal_columns:
                    if seal_col in source_col_indices:
                        source_col_idx = source_col_indices[seal_col]
                        value = ws_source.cell(row=source_row, column=source_col_idx).value
                        value = self.get_non_empty_value(value)
                        
                        # Special handling for formats with " / " separator
                        # Both Rodman and UnitList use "seal / type / origin" format
                        if value and '/' in str(value) and file_format in ['RODMAN', 'RODMAN_CONVERTED', 'UNITLIST']:
                            # Extract seal number before first " / "
                            value = str(value).split('/')[0].strip()
                            if not value:
                                continue
                        
                        if value is not None and value not in seals_added:
                            seals_list.append(value)
                            seals_added.add(value)
                
                # Write seals to template
                for i, seal_value in enumerate(seals_list):
                    if i < 4:
                        # Seals 1-4: one per column
                        template_col_idx = template_cols[f'Seal {i+1}']
                        ws_template.cell(row=output_row, column=template_col_idx).value = seal_value
                    else:
                        # Seal 5 and beyond: combine with separator
                        template_col_idx = template_cols['Seal 5']
                        if i == 4:
                            ws_template.cell(row=output_row, column=template_col_idx).value = seal_value
                        else:
                            current = ws_template.cell(row=output_row, column=template_col_idx).value
                            new_value = f"{current}  ;  {seal_value}"
                            ws_template.cell(row=output_row, column=template_col_idx).value = new_value
                
                total_seals += len(seals_list)
                num_seals = len(seals_list)
                if num_seals > 0:
                    if num_seals not in seal_distribution:
                        seal_distribution[num_seals] = 0
                    seal_distribution[num_seals] += 1
                
                output_row += 1
                
                # Progress - show containers and seals
                if containers_found % 100 == 0:
                    log(f"   ⏳ {containers_found} containers, {total_seals} seals...")
            
            # Don't save if no containers found (when filtering)
            if containers_found == 0:
                log(f"\n⚠️  No containers found matching the filter criteria")
                log(f"   Output file NOT created")
                return {
                    'containers': 0,
                    'seals': 0,
                    'output': None,
                    'distribution': {}
                }
            
            # Save output
            log(f"\n💾 Saving: {Path(output_path).name}")
            log(f"   Location: {Path(output_path).parent}")
            
            wb_template.save(output_path)
            wb_template.close()
            
            # Re-open and re-save to normalize (fixes Excel NaN display issue)
            log(f"🔄 Normalizing file...")
            wb_normalized = openpyxl.load_workbook(output_path)
            wb_normalized.save(output_path)
            wb_normalized.close()
            
            # Statistics
            log(f"\n✅ CONVERSION COMPLETED!")
            log(f"\n📊 STATISTICS:")
            log(f"   • Containers: {containers_found} ✓")
            if containers_skipped > 0:
                log(f"   • Skipped (filtered): {containers_skipped}")
            log(f"   • Seals: {total_seals} ✓")
            
            log(f"\n💾 OUTPUT:")
            log(f"   {output_path}\n")
            
            return {
                'containers': containers_found,
                'seals': total_seals,
                'output': output_path,
                'distribution': seal_distribution,
                'pol': self.get_pol_value(ws_template)  # Return POL for naming
            }
            
        except Exception as e:
            log(f"\n❌ ERROR: {str(e)}")
            import traceback
            log(traceback.format_exc())
            return None